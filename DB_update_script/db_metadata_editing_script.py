"""
# Metadata editing script for ArcGIS Pro
# This script is designed to modify metadata in a file geodatabase by importing a template XML file.
# There is no undo functionality, so use with care! - it overwrites existing metadata.
# As a safeguard, it exports existing metadata to an XML file before modifying it.
# It exports existing metadata to an XML file and creates an Excel spreadsheet listing dataset titles and descriptions

# modified from older examples at https://github.com/ucd-cws/arcpy_metadata

********************
Script Use:

This script is supposed to be run within an ArcGIS Pro script tool (in a toolbox)
it expects two parameters: (1) the workspace to run against (2) a template Metadata file in XML format

- This script modifies (overwrites) the metadata in the target workspace (!) - USE WITH CARE!!

- This script exports existing metadata into XML Metadata files (in the APPDATA/local/Esri/ArcGISPro directory)

- This script creates an XLSX output file that lists the titles and descriptions of all datasets in the GDB (also in Local Appdata dir)

PURPOSE:
The code in this notebook can be used to import the contents of a template Metadata XML file into all feature classes in a file GDB

Specify locations for workspace (Geodatabase) and Template (XML file) first
The template can be modified according to https://pro.arcgis.com/en/pro-app/latest/help/metadata/best-practices-for-editing-metadata.htm

Known LIMITATIONS: Currently the XML file contents are not imported to Feature Datasets (folders) and the main level of the GDB itself.


"""

import datetime
import os

import arcpy

# import module for Excel sheets
import openpyxl
from arcpy import metadata as md

###### Configure a log file location

datetime_object = datetime.datetime.now()
logfileDate = datetime_object.strftime("%d_%m_%Y_%H%M%S")
# output file
appd = os.getenv("LOCALAPPDATA")
arcpy.AddMessage(f"LOCALAPPDATA: {appd}")

filename = os.path.join(appd, "ESRI", "ArcGISPro", f"GDBcontents_{logfileDate}.txt")
x = os.path.join(appd, "ESRI", "ArcGISPro", f"GDBcontents_{logfileDate}.xlsx")

arcpy.AddMessage(f"Logfile name: {filename}")


# make sure we also capture the execution date in the filename of the exported metadata
fileDateMD = datetime_object.strftime("%d_%m_%Y_%H%M%S")

# Call a Workbook() function of openpyxl to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet from the active attribute
sheet = wb.active


def write_spreadsheet(spreadsheetFile, needProjectName=None, projectName=None):
    """Cell objects also have row, column and coordinate attributes that provide
    location information for the cell.

    Note: The first row or column integer is 1, not 0. Cell object is created by
    using sheet object's cell() method."""
    c1 = sheet.cell(row=1, column=1)

    # writing values to cells
    c1.value = "Name"

    c2 = sheet.cell(row=1, column=2)
    c2.value = "Title"

    c3 = sheet.cell(row=1, column=3)

    # writing values to cells
    c3.value = "Summary"

    c4 = sheet.cell(row=1, column=4)
    c4.value = "Description"

    c5 = sheet.cell(row=1, column=5)

    # writing values to cells
    c5.value = "Tags"

    c6 = sheet.cell(row=1, column=6)
    c6.value = "Credits"

    c7 = sheet.cell(row=1, column=7)

    # writing values to cells
    c7.value = "MaxScale"

    c8 = sheet.cell(row=1, column=8)
    c8.value = "MinScale"

    c9 = sheet.cell(row=1, column=9)
    c8.value = "Access restrictions"

    # # Iterate over Feature Classes and write their MD content (title, description) into the log file
    datasets = arcpy.ListDatasets()
    # Find the total count of Feature Classes in list
    fc_count = len(datasets)

    # Set the progress indicator
    arcpy.SetProgressor(
        "step", "Modify Metadata for Feature Classes in geodatabase...", 0, fc_count, 1
    )

    i = 2
    datasets = [""] + datasets if datasets is not None else []
    for ds in datasets:
        # read the contents of the GDB, print to screen, write to log file and export MD to an XML file output:
        for fc in arcpy.ListFeatureClasses("*", "All", feature_dataset=ds):
            # Update the progressor label for current shapefile
            arcpy.SetProgressorLabel(f"Loading {fc}...")

            # Get metadata for current feature:
            metadata = md.Metadata(
                fc
            )  # currently supports Shapefiles, FeatureClasses, RasterDatasets and Layers
            title = metadata.title
            abstract = metadata.description
            tags = metadata.tags
            credits = metadata.credits
            summary = metadata.summary
            use_restrictions = metadata.accessConstraints
            mxScale = metadata.maxScale
            mnScale = metadata.minScale
            # find cell position on spreadsheet to write to:
            nameCell = sheet.cell(row=i, column=1)
            titleCell = sheet.cell(row=i, column=2)
            summaryCell = sheet.cell(row=i, column=3)
            descriptionCell = sheet.cell(row=i, column=4)
            tagsCell = sheet.cell(row=i, column=5)
            creditsCell = sheet.cell(row=i, column=6)
            scale_mxCell = sheet.cell(row=i, column=7)
            scale_mnCell = sheet.cell(row=i, column=8)
            use_restrictionsCell = sheet.cell(row=i, column=9)
            # transfer metadata to spreadsheet
            if abstract != "None":
                try:
                    arcpy.AddMessage(
                        f"Title {title} has following abstract:\n{abstract}\n"
                    )
                    # f.write(f"Title {title} has following abstract:\n{abstract}\n")

                    titleCell.value = f"{title}"
                    descriptionCell.value = f"{abstract}"

                except:
                    arcpy.AddWarning(
                        f"Error while exporting content for {title} into spreadsheet"
                    )
            else:
                try:
                    arcpy.AddMessage(f"Title {title} has no abstract")
                    # f.write(f"Title {title} has no abstract")
                    titleCell.value = f"{title}"
                    descriptionCell.value = "No description"
                except:
                    arcpy.AddWarning(
                        f"Error while printing content to XLSX for {title}"
                    )
            try:
                arcpy.AddMessage(
                    f"Adding other metadata values for  {fc} with title {title}\n"
                )
                nameCell.value = f"{fc}"  # = sheet.cell(row = i , column = 1)
                use_restrictionsCell.value = f"{use_restrictions}"
                summaryCell.value = f"{summary}"  # = sheet.cell(row = i, column = 3)
                tagsCell.value = f"{tags}"  # = sheet.cell(row = i, column = 5)
                creditsCell.value = f"{credits}"  # = sheet.cell(row = i, column = 6)
                scale_mxCell.value = f"{mxScale}"  # = sheet.cell(row = i, column = 7)
                scale_mnCell.value = f"{mnScale}"  # = sheet.cell(row = i, column = 8)
            except:
                arcpy.AddWarning(
                    f"Error while adding other metadata values for  {fc}  into spreadsheet"
                )

            # Update the progressor position
            arcpy.SetProgressorPosition()
            i += 1

            # # f.write(f"**********\n  Feature Dataset: {ds} **********\n")
            arcpy.AddMessage(f"Feature class {fc}")
            catalogueMetadataContent(
                fc,
                templateImportFile,
                fileDateMD,
                metadata,
                needProjectName,
                projectName,
            )

        arcpy.ResetProgressor()

    # Anytime you modify the Workbook object or its sheets and cells, the spreadsheet
    # file will not be saved until you call the save() workbook method.
    wb.save(spreadsheetFile)
    arcpy.AddMessage(f"Spreadsheet saved to {spreadsheetFile}")


def catalogueMetadataContent(FC, TIF, FDMD, MD, needprojn, projn):
    # read the contents of the GDB, export MD to an XML file output for each FC:
    arcpy.AddMessage(f"\nTrying to export MD for {FC} to XML file:\n")

    # export a backup copy of the Metadata into an XML file as a safeguard:
    export_19115_3_path = os.path.join(
        appd, "ESRI", "ArcGISPro", f"{FC}_{FDMD}_19115_3.xml"
    )
    try:
        MD.exportMetadata(
            export_19115_3_path, metadata_removal_option="REMOVE_MACHINE_NAMES"
        )
        arcpy.AddMessage(f"Successfully exported MD to file {export_19115_3_path}.")

    except:
        arcpy.AddError(f"Failed to export MD for {FC} to {export_19115_3_path}\n\n")

    arcpy.AddMessage(f"Trying to import MD from template {TIF} to {FC}\n")
    # Import and overwrite metadata in GDB with those from a template:
    try:
        # Get metadata content from a metadata template XML file
        src_item_md = md.Metadata(TIF)

        # if metadata like title or description have previously been modified, preserve them, instead of overwriting.
        if needprojn == "true":
            arcpy.AddWarning(f"Found needProjectName {projn} is {needprojn}\n")
            prev_title = f"{projn} " + f"{MD.title}"  # = 'My Title'
            prev_tags = f"{projn}, " + f"{MD.tags}"  # = 'Tag1, Tag2'
            prev_summary = f"{MD.summary}.\nProject: {projn}"  # = 'My Summary'
            prev_description = (
                f"{MD.description} \nProject: {projn}"  # = 'My Description'
            )
            prev_credits = MD.credits  # = 'My Credits'
            prev_accessconstraints = MD.accessConstraints  # = 'My Access Constraints'
        else:
            prev_title = MD.title  # = 'My Title'
            prev_tags = MD.tags  # = 'Tag1, Tag2'
            prev_summary = MD.summary  # = 'My Summary'
            prev_description = MD.description  # = 'My Description'
            prev_credits = MD.credits  # = 'My Credits'
            prev_accessconstraints = MD.accessConstraints  # = 'My Access Constraints'

        if not MD.isReadOnly:
            # Copy the template's content to a feature class in the file geodatabase
            MD.copy(src_item_md)
            arcpy.AddMessage(f"Successfully imported MD from {TIF} into {FC}")

    except:
        arcpy.AddError(f"Failed to import MD from {TIF} to {FC}")

    arcpy.AddMessage(f"Trying to add old MD back into {FC} after importing from file\n")
    try:
        # Restore previous metadata by adding them to the newly copied content:
        if prev_title != "":
            arcpy.AddMessage(f"Adding old MD title {prev_title} back")
            MD.title = f"{prev_title}" + f"{MD.title}"
        else:
            arcpy.AddWarning(
                f"Nothing to add (old MD title '{prev_title}' for {FC}) \n\n"
            )

        if prev_description != "None":
            MD.description = f"{prev_description};\n" + f"{MD.description}"
        else:
            arcpy.AddWarning(
                f"Nothing to add (old MD description '{prev_description}' for {FC}) \n\n"
            )

        if prev_credits != "None":
            MD.credits = f"{prev_credits}/ " + f"{MD.credits}"
        else:
            arcpy.AddWarning(
                f"Nothing to add (old MD credits '{prev_credits}' for {FC}) \n\n"
            )

        if prev_summary != "None":
            MD.summary = f"{prev_summary}; " + f"{MD.summary}"
        else:
            arcpy.AddWarning(
                f"Nothing to add (old MD summary '{prev_summary}' for {FC}) \n\n"
            )

        if prev_tags != "None":
            MD.tags = f"{prev_tags}," + f"{MD.tags}"
        else:
            arcpy.AddWarning(
                f"Nothing to add (old MD tags '{prev_tags}' for {FC}) \n\n"
            )

        if prev_accessconstraints != "None":
            MD.accessConstraints = (
                f"{prev_accessconstraints}," + f"{MD.accessConstraints}"
            )
        else:
            arcpy.AddWarning(
                f"Nothing to add (old MD access constraints '{prev_accessconstraints}' for {FC}) \n\n"
            )

        MD.save()

    except:
        arcpy.AddError(
            f"Failed to re-add old MD for {FC} \n\nRolling back to old MD.\n\n"
        )


if __name__ == "__main__":
    # Configure Geodatabase:
    arcpy.env.workspace = arcpy.GetParameterAsText(0)

    # Configure Metadata Template file that will be imported:
    templateImportFile = arcpy.GetParameterAsText(1)

    needProjectName = arcpy.GetParameterAsText(3)
    arcpy.AddMessage(f"Need project name is: {needProjectName}")

    # Configure Metadata Template file that will be imported:
    projectName = arcpy.GetParameterAsText(4)
    arcpy.AddMessage(f"Project name is: {projectName}")

    if needProjectName:
        write_spreadsheet(x, needProjectName, projectName)
    else:
        write_spreadsheet(x)

        # return the output to the Geoprocessing console in ArcGIS Pro
        arcpy.SetParameterAsText(2, f"{x}")
